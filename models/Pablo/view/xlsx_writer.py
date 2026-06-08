# =============================================================================
# view/xlsx_writer.py
# Write the Michelin evaluation table to an Excel (.xlsx) file.
#
# Format: nuevas_normas.pdf (June 2026).
# One worksheet per image named "Foto 1", "Foto 2", "Foto 3".
# Each sheet has the exact two-section table Michelin will compare against
# their ground truth.
#
# Reto 1  — 9 fixed X positions (80…720 mm), columns 1–9, column 10 empty.
#   DA, DB, LA, LB
#
# Reto 2  — 10 fixed Y offsets (5…230 mm from band-cut top), columns 1–10.
#   SA1, SA2, YSA, SB1, SB2, YSB
# =============================================================================

from __future__ import annotations

import os
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from model.reto_measurements import Reto1Row, Reto2Row
from config.config import RETO1_X_MM, RETO2_Y_OFFSETS_MM


# ── Colour palette ────────────────────────────────────────────────────────────
_BLUE_HEADER  = PatternFill("solid", fgColor="1F4E79")   # dark blue
_BLUE_LIGHT   = PatternFill("solid", fgColor="BDD7EE")   # light blue (Reto 1 sub-header)
_GREEN_LIGHT  = PatternFill("solid", fgColor="E2EFDA")   # light green (Reto 2 sub-header)
_GREY         = PatternFill("solid", fgColor="D9D9D9")   # row label
_WHITE        = PatternFill("solid", fgColor="FFFFFF")

_WHITE_FONT   = Font(bold=True, color="FFFFFF", size=11)
_BOLD         = Font(bold=True, size=10)
_NORMAL       = Font(size=10)
_SMALL        = Font(size=9, italic=True)

_THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
_CTR = Alignment(horizontal="center", vertical="center", wrap_text=False)
_LEFT = Alignment(horizontal="left",  vertical="center", wrap_text=True)

# Number of data columns in the shared table
_N_COLS = 10   # max(9 for Reto 1, 10 for Reto 2)


def _c(ws, row: int, col: int, value=None,
        fill=None, font=None, align=None, border=True):
    """Set a cell, applying optional formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:   cell.fill   = fill
    if font:   cell.font   = font
    if align:  cell.alignment = align
    if border: cell.border = _THIN
    return cell


def _num(val: Optional[float]) -> Optional[float]:
    """Return a rounded float or None (empty cell) for missing values."""
    if val is None:
        return None
    return round(float(val), 2)


# ── Public API ────────────────────────────────────────────────────────────────

def write_reto_xlsx(path: str, results: list[dict]) -> None:
    """Write one xlsx with one sheet per image (Foto 1, Foto 2, Foto 3…).

    ``results`` list of dicts with keys:
        image_name   str
        reto1_rows   list[Reto1Row]   (9 items)
        reto2_rows   list[Reto2Row]   (10 items)
        qr_ok        bool
        n_edges      int
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for foto_num, rec in enumerate(results, start=1):
        ws = wb.create_sheet(title=f"Foto {foto_num}")
        _write_foto_sheet(ws, rec, foto_num)

    wb.save(path)
    print(f"[XlsxWriter] Saved: {path}")


def _write_foto_sheet(ws, rec: dict, foto_num: int) -> None:
    reto1: list[Reto1Row] = rec.get("reto1_rows", [])
    reto2: list[Reto2Row] = rec.get("reto2_rows", [])

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 16   # metric label
    ws.column_dimensions["B"].width = 18   # first data col (sometimes description)
    for i in range(1, _N_COLS + 1):
        ws.column_dimensions[get_column_letter(1 + i)].width = 9

    row = 1

    # ── Section title: Foto N ─────────────────────────────────────────────────
    ws.row_dimensions[row].height = 22
    img_name = rec.get("image_name", f"imagen_{foto_num}")
    qr_ok    = "QR OK" if rec.get("qr_ok") else "SIN QR"
    _c(ws, row, 1, f"Foto {foto_num}  —  {img_name}  [{qr_ok}]",
       fill=_BLUE_HEADER, font=_WHITE_FONT, align=_LEFT)
    # Merge title across all columns
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row,   end_column=1 + _N_COLS,
    )
    row += 1

    # ══════════════════════════════════════════════════════════════════════════
    # RETO 1
    # ══════════════════════════════════════════════════════════════════════════
    row = _write_reto1_section(ws, row, reto1)

    row += 1  # blank separator between sections

    # ══════════════════════════════════════════════════════════════════════════
    # RETO 2
    # ══════════════════════════════════════════════════════════════════════════
    row = _write_reto2_section(ws, row, reto2)

    row += 2  # gap before legend

    # ── Legend ────────────────────────────────────────────────────────────────
    _write_legend(ws, row)


# ── Reto 1 section ────────────────────────────────────────────────────────────

def _write_reto1_section(ws, start_row: int, rows: list[Reto1Row]) -> int:
    r = start_row

    # Column index header:  blank | 1 | 2 | … | 9 | (10 empty)
    ws.row_dimensions[r].height = 18
    _c(ws, r, 1, "", fill=_BLUE_LIGHT, font=_BOLD, align=_CTR)
    for j in range(1, _N_COLS + 1):
        label = str(j) if j <= 9 else ""
        _c(ws, r, 1 + j, label, fill=_BLUE_LIGHT, font=_BOLD, align=_CTR)
    r += 1

    # Distance sub-header: "Distancia a borde (mm)" | 80 | 160 | … | 720 | ""
    ws.row_dimensions[r].height = 16
    _c(ws, r, 1, "Distancia a borde (mm)", fill=_BLUE_LIGHT, font=_BOLD, align=_CTR)
    for j, x in enumerate(RETO1_X_MM):          # 9 values
        _c(ws, r, 2 + j, x, fill=_BLUE_LIGHT, font=_BOLD, align=_CTR)
    _c(ws, r, 2 + 9, "", fill=_BLUE_LIGHT, font=_BOLD, align=_CTR)   # col 10 empty
    r += 1

    # Data rows
    reto1_attrs = ["DA", "DB", "LA", "LB"]
    for attr in reto1_attrs:
        ws.row_dimensions[r].height = 16
        _c(ws, r, 1, attr, fill=_GREY, font=_BOLD, align=_CTR)
        for j, row_data in enumerate(rows):          # up to 9 items
            val = _num(getattr(row_data, attr, None))
            _c(ws, r, 2 + j, val, fill=_WHITE, font=_NORMAL, align=_CTR)
        # fill remaining columns up to _N_COLS with empty cells
        for j in range(len(rows), _N_COLS):
            _c(ws, r, 2 + j, None, fill=_WHITE, font=_NORMAL, align=_CTR)
        r += 1

    return r


# ── Reto 2 section ────────────────────────────────────────────────────────────

def _write_reto2_section(ws, start_row: int, rows: list[Reto2Row]) -> int:
    r = start_row

    # Column index header:  blank | 1 | 2 | … | 10
    ws.row_dimensions[r].height = 18
    _c(ws, r, 1, "", fill=_GREEN_LIGHT, font=_BOLD, align=_CTR)
    for j in range(1, _N_COLS + 1):
        _c(ws, r, 1 + j, str(j), fill=_GREEN_LIGHT, font=_BOLD, align=_CTR)
    r += 1

    # Distance sub-header: "Distancia a borde producto (mm)" | 5 | 30 | … | 230
    ws.row_dimensions[r].height = 16
    _c(ws, r, 1, "Distancia a borde producto (mm)",
       fill=_GREEN_LIGHT, font=_BOLD, align=_CTR)
    for j, y in enumerate(RETO2_Y_OFFSETS_MM):  # 10 values
        _c(ws, r, 2 + j, y, fill=_GREEN_LIGHT, font=_BOLD, align=_CTR)
    r += 1

    # Data rows
    reto2_attrs = ["SA1", "SA2", "YSA", "SB1", "SB2", "YSB"]
    for attr in reto2_attrs:
        ws.row_dimensions[r].height = 16
        _c(ws, r, 1, attr, fill=_GREY, font=_BOLD, align=_CTR)
        for j, row_data in enumerate(rows):          # up to 10 items
            val = _num(getattr(row_data, attr, None))
            _c(ws, r, 2 + j, val, fill=_WHITE, font=_NORMAL, align=_CTR)
        for j in range(len(rows), _N_COLS):
            _c(ws, r, 2 + j, None, fill=_WHITE, font=_NORMAL, align=_CTR)
        r += 1

    return r


# ── Legend ────────────────────────────────────────────────────────────────────

_LEGEND = [
    ("DA",  "Distancia a borde mesa — banda A borde mas proximo (mm)"),
    ("DB",  "Distancia a borde mesa — banda B borde mas proximo (mm)"),
    ("LA",  "Anchura de banda A (mm)"),
    ("LB",  "Anchura de banda B (mm)"),
    ("SA1", "Distancia entre bordes superiores de soldadura en Banda A (mm)"),
    ("SA2", "Distancia entre borde superior e inferior (entre negros) de soldadura en Banda A (mm)"),
    ("YSA", "Distancia en eje Y al borde de la mesa de borde soldadura mas proximo en Banda A (mm)"),
    ("SB1", "Distancia entre bordes superiores de soldadura en Banda B (mm)"),
    ("SB2", "Distancia entre borde superior e inferior (entre negros) de soldadura en Banda B (mm)"),
    ("YSB", "Distancia en eje Y al borde de la mesa de borde soldadura mas proximo en Banda B (mm)"),
]


def _write_legend(ws, start_row: int) -> None:
    r = start_row
    _c(ws, r, 1, "Leyenda", fill=_GREY, font=_BOLD, align=_LEFT, border=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r,
                   end_column=1 + _N_COLS)
    r += 1
    for abbr, desc in _LEGEND:
        _c(ws, r, 1, abbr,  fill=None, font=_BOLD,   align=_CTR,  border=False)
        cell = ws.cell(row=r, column=2, value=desc)
        cell.font = _SMALL
        cell.alignment = _LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r,
                       end_column=1 + _N_COLS)
        r += 1


# ── Template-filling writer (Michelin official format) ────────────────────────

def write_reto_xlsx_from_template(
    template_path: str,
    output_path: str,
    results: list[dict],
) -> None:
    """Fill the official Michelin xlsx template with pipeline measurements.

    Preserves all original template formatting and descriptions.
    Template layout: single sheet "Hoja1", one Foto section per block of 13
    rows starting at row 5, with a blank row between sections.

    Layout per Foto section (row offsets from section start):
      +0  title row ("Foto N")  + column index 1-10
      +1  "Distancia a borde"   + 80,160,…,720
      +2  DA   (cols 2-10, 9 values)
      +3  DB
      +4  LA
      +5  LB
      +6  "Distancia a borde producto" + 5,30,…,230
      +7  SA1  (cols 2-11, 10 values)
      +8  SA2
      +9  YSA  (same value repeated across 10 cols)
      +10 SB1
      +11 SB2
      +12 YSB  (same value repeated across 10 cols)
    """
    import shutil
    shutil.copy2(template_path, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Hoja1"]

    # Exact starting rows of each Foto section in the template
    # (verified by inspecting the template file directly)
    FOTO_STARTS = [5, 21, 36]

    # Row offsets within each foto section for Reto 1 metrics
    RETO1_ROW_OFFSETS = {"DA": 2, "DB": 3, "LA": 4, "LB": 5}
    # Row offsets within each foto section for Reto 2 metrics
    RETO2_ROW_OFFSETS = {"SA1": 7, "SA2": 8, "YSA": 9,
                         "SB1": 10, "SB2": 11, "YSB": 12}

    for foto_idx, rec in enumerate(results[:3]):
        foto_start = FOTO_STARTS[foto_idx]

        reto1: list[Reto1Row] = rec.get("reto1_rows", [])
        reto2: list[Reto2Row] = rec.get("reto2_rows", [])

        # ── Reto 1: DA, DB, LA, LB — 9 data columns (cols 2-10) ─────────────
        for attr, r_off in RETO1_ROW_OFFSETS.items():
            row = foto_start + r_off
            for j, row_data in enumerate(reto1[:9]):
                ws.cell(row=row, column=2 + j).value = _num(
                    getattr(row_data, attr, None)
                )

        # ── Reto 2: SA1,SA2,YSA,SB1,SB2,YSB — 10 data cols (cols 2-11) ─────
        for attr, r_off in RETO2_ROW_OFFSETS.items():
            row = foto_start + r_off
            for j, row_data in enumerate(reto2[:10]):
                ws.cell(row=row, column=2 + j).value = _num(
                    getattr(row_data, attr, None)
                )

    wb.save(output_path)
    print(f"[XlsxWriter] Template saved: {output_path}")
